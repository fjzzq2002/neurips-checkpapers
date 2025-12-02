"""
Rebuild example_workbook.xlsx from the JSON sources.

Columns are ordered to match the example sheet, with Abstract placed after Summary (CN).
Header row is bold. Rows are sorted by Session, then Weighted Score (interp + understanding + safety*2.5, desc).
Row fills:
  - Safety?=1 only: green-ish (FFD9EAD3)
  - MI?=1 only: orange-ish (FFFCE5CD)
  - Both: pink-ish (FFEAD1DC)
"""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parents[1]
ANNOTATIONS = ROOT / "data/analysis/neurips_2025_annotated.jsonl"
SCRAPED = ROOT / "data/neurips_2025_san_diego_posters.json"
OUTPUT = ROOT / "example_workbook.xlsx"


def load_scraped() -> Dict[str, Dict[str, Any]]:
    payload = json.loads(SCRAPED.read_text(encoding="utf-8"))
    records = payload.get("records", [])
    authors = {}
    abstracts = {}
    for rec in records:
        rid = rec.get("id")
        if rid is None:
            continue
        key = str(int(rid)) if isinstance(rid, (int, float)) else str(rid)
        authors[key] = rec.get("authors") or []
        abstracts[key] = rec.get("abstract") or ""
    return {"authors": authors, "abstracts": abstracts}


def fmt_authors(auths: Any) -> str:
    names = []
    for a in auths or []:
        if isinstance(a, str):
            names.append(a)
        elif isinstance(a, dict):
            nm = a.get("fullname") or a.get("name")
            if nm:
                names.append(str(nm))
    return ", ".join(names)


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if not isinstance(value, str):
        value = str(value)
    return ILLEGAL_CHARACTERS_RE.sub("", value)


def clean_session(name: str) -> str:
    if not name:
        return ""
    return name.replace("San Diego ", "")


def poster_number(session_name: str) -> str:
    parts = session_name.replace("#", " ").split()
    for token in parts:
        if token.isdigit():
            return token
    return ""


def load_rows(authors_map: Dict[str, Any], abstracts_map: Dict[str, str]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    with ANNOTATIONS.open("r", encoding="utf-8") as handle:
        for line in handle:
            line = line.strip()
            if not line:
                continue
            rec = json.loads(line)
            rid_raw = rec.get("record_id")
            rid_key = str(int(rid_raw)) if isinstance(rid_raw, (int, float)) else str(rid_raw)
            analysis = rec.get("analysis") or {}
            scoring = analysis.get("scoring") or {}
            cat = analysis.get("category") or {}
            verdict = analysis.get("verdict") or {}
            sessions = rec.get("sessions") or []

            session_names = []
            poster_entry = None
            for s in sessions:
                nm = clean_session(s.get("name") or "")
                kind = (s.get("kind") or "").lower()
                if kind == "poster" or "poster session" in nm.lower():
                    poster_entry = s
                if nm:
                    session_names.append(nm)
            session_str = " | ".join(session_names)

            poster_value = ""
            if poster_entry:
                nm = clean_session(poster_entry.get("name") or "")
                num = poster_number(nm)
                pos = poster_entry.get("poster_position") or ""
                poster_value = " ".join([p for p in (num, pos) if p])

            interp = int(scoring.get("interpretability") or 0)
            und = int(scoring.get("understanding") or 0)
            saf = float(scoring.get("safety") or 0)
            tech = scoring.get("technicality")
            surpr = scoring.get("surprisal")
            weighted = interp + und + saf * 2.5

            rows.append(
                {
                    "ID": int(float(rid_raw)) if rid_raw is not None else None,
                    "Title": clean_text(rec.get("title")),
                    "Authors": clean_text(fmt_authors(authors_map.get(rid_key))),
                    "Decision": (rec.get("presentation") or rec.get("decision") or "").lower(),
                    "Topic": clean_text(rec.get("topic")),
                    "Session": clean_text(session_str),
                    "Poster #": clean_text(poster_value),
                    "Safety?": verdict.get("is_ai_safety"),
                    "MI?": verdict.get("is_mech_interp"),
                    "Link": rec.get("virtualsite_url") or rec.get("paper_url"),
                    "Summary": clean_text(analysis.get("summary", "")),
                    "Summary (CN)": clean_text(analysis.get("summary_cn", "")),
                    "Abstract": clean_text(abstracts_map.get(rid_key, "")),
                    "Focus": cat.get("primary_focus"),
                    "Failure Mode": cat.get("failure_mode_addressed"),
                    "interpretability": interp,
                    "understanding": und,
                    "safety": saf,
                    "technicality": tech,
                    "surprisal": surpr,
                    "Keywords": clean_text(analysis.get("keywords", "")),
                    "Weighted Score": weighted,
                }
            )
    return rows


def build_workbook(rows: List[Dict[str, Any]], output_path: Path) -> None:
    # Sorting: Session asc, then Weighted Score desc, then safety/interp/understanding desc
    rows.sort(
        key=lambda r: (
            r["Session"],
            -r["Weighted Score"],
            -(r["safety"] or 0),
            -(r["interpretability"] or 0),
            -(r["understanding"] or 0),
        )
    )

    headers = [
        "ID",
        "Title",
        "Authors",
        "Decision",
        "Topic",
        "Session",
        "Poster #",
        "Safety?",
        "MI?",
        "Link",
        "Summary",
        "Summary (CN)",
        "Abstract",
        "Focus",
        "Failure Mode",
        "interpretability",
        "understanding",
        "safety",
        "technicality",
        "surprisal",
        "Keywords",
        "Weighted Score",
    ]

    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    bold_font = Font(bold=True)
    for col_idx in range(1, len(headers) + 1):
        ws.cell(row=1, column=col_idx).font = bold_font

    fill_safety = PatternFill(fill_type="solid", fgColor="FFD9EAD3")
    fill_mi = PatternFill(fill_type="solid", fgColor="FFFCE5CD")
    fill_both = PatternFill(fill_type="solid", fgColor="FFEAD1DC")

    for row in rows:
        ws.append([row.get(h) for h in headers])
        safety_flag = row.get("Safety?") == 1 or row.get("Safety?") is True
        mi_flag = row.get("MI?") == 1 or row.get("MI?") is True
        decision = (row.get("Decision") or "").lower()
        fill = None
        if safety_flag and mi_flag:
            fill = fill_both
        elif safety_flag:
            fill = fill_safety
        elif mi_flag:
            fill = fill_mi
        if fill:
            for c in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=c).fill = fill
        if "oral" in decision or "spotlight" in decision:
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> None:
    scraped = load_scraped()
    rows = load_rows(scraped["authors"], scraped["abstracts"])
    build_workbook(rows, OUTPUT)
    print(f"Wrote {OUTPUT}")


if __name__ == "__main__":
    main()
